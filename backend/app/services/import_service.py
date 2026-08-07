import csv
import io

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.prospect import Prospect


def clean_value(value: str | None) -> str | None:
    if value is None:
        return None

    value = value.strip()

    return value if value else None


def import_prospects_from_csv(
    db: Session,
    content: bytes,
) -> dict[str, int]:
    text = content.decode("utf-8-sig")

    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    duplicates = 0
    ignored = 0

    for row in reader:
        company_name = clean_value(
            row.get("company_name")
        )

        if not company_name:
            ignored += 1
            continue

        public_email = clean_value(
            row.get("public_email")
        )

        website = clean_value(
            row.get("website")
        )

        duplicate = None

        if public_email:
            duplicate = db.scalar(
                select(Prospect).where(
                    Prospect.public_email == public_email
                )
            )

        if duplicate is None and website:
            duplicate = db.scalar(
                select(Prospect).where(
                    Prospect.website == website
                )
            )

        if duplicate is not None:
            duplicates += 1
            continue

        try:
            priority = int(
                clean_value(row.get("priority")) or 3
            )
        except ValueError:
            priority = 3

        priority = max(1, min(priority, 5))

        try:
            score = float(
                clean_value(row.get("score")) or 0
            )
        except ValueError:
            score = 0

        score = max(0, min(score, 100))

        prospect = Prospect(
            company_name=company_name,
            country=clean_value(row.get("country")),
            city=clean_value(row.get("city")),
            website=website,
            linkedin=clean_value(row.get("linkedin")),
            public_email=public_email,
            public_phone=clean_value(
                row.get("public_phone")
            ),
            industry=clean_value(row.get("industry")),
            priority=priority,
            status=clean_value(row.get("status"))
            or "À contacter",
            score=score,
        )

        db.add(prospect)
        imported += 1

    db.commit()

    return {
        "imported": imported,
        "duplicates": duplicates,
        "ignored": ignored,
    }
import csv
import io
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.prospect import Prospect


def clean_value(value: Any) -> str | None:
    if value is None:
        return None

    value = str(value).strip()

    return value if value else None


def parse_priority(value: Any) -> int:
    try:
        priority = int(clean_value(value) or 3)
    except (ValueError, TypeError):
        priority = 3

    return max(1, min(priority, 5))


def parse_score(value: Any) -> float:
    try:
        score = float(clean_value(value) or 0)
    except (ValueError, TypeError):
        score = 0

    return max(0, min(score, 100))


def prospect_exists(
    db: Session,
    public_email: str | None,
    website: str | None,
) -> bool:
    if public_email:
        duplicate = db.scalar(
            select(Prospect).where(
                Prospect.public_email == public_email
            )
        )

        if duplicate is not None:
            return True

    if website:
        duplicate = db.scalar(
            select(Prospect).where(
                Prospect.website == website
            )
        )

        if duplicate is not None:
            return True

    return False


def import_rows(
    db: Session,
    rows: list[dict[str, Any]],
) -> dict[str, int]:
    imported = 0
    duplicates = 0
    ignored = 0

    for row in rows:
        company_name = clean_value(
            row.get("company_name")
        )

        if not company_name:
            ignored += 1
            continue

        public_email = clean_value(
            row.get("public_email")
        )

        website = clean_value(
            row.get("website")
        )

        if prospect_exists(
            db,
            public_email,
            website,
        ):
            duplicates += 1
            continue

        prospect = Prospect(
            company_name=company_name,
            country=clean_value(
                row.get("country")
            ),
            city=clean_value(
                row.get("city")
            ),
            website=website,
            linkedin=clean_value(
                row.get("linkedin")
            ),
            public_email=public_email,
            public_phone=clean_value(
                row.get("public_phone")
            ),
            industry=clean_value(
                row.get("industry")
            ),
            priority=parse_priority(
                row.get("priority")
            ),
            status=(
                clean_value(row.get("status"))
                or "À contacter"
            ),
            score=parse_score(
                row.get("score")
            ),
        )

        db.add(prospect)

        # Permet aussi de détecter les doublons
        # présents plusieurs fois dans le même fichier.
        db.flush()

        imported += 1

    db.commit()

    return {
        "imported": imported,
        "duplicates": duplicates,
        "ignored": ignored,
    }


def import_prospects_from_csv(
    db: Session,
    content: bytes,
) -> dict[str, int]:
    text = content.decode("utf-8-sig")

    reader = csv.DictReader(
        io.StringIO(text)
    )

    rows = [
        dict(row)
        for row in reader
    ]

    return import_rows(
        db,
        rows,
    )


def import_prospects_from_xlsx(
    db: Session,
    content: bytes,
) -> dict[str, int]:
    workbook = load_workbook(
        filename=io.BytesIO(content),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    row_iterator = worksheet.iter_rows(
        values_only=True
    )

    try:
        header_row = next(row_iterator)
    except StopIteration:
        return {
            "imported": 0,
            "duplicates": 0,
            "ignored": 0,
        }

    headers = [
        clean_value(value)
        for value in header_row
    ]

    rows: list[dict[str, Any]] = []

    for values in row_iterator:
        row: dict[str, Any] = {}

        for index, header in enumerate(headers):
            if not header:
                continue

            value = (
                values[index]
                if index < len(values)
                else None
            )

            row[header] = value

        rows.append(row)

    workbook.close()

    return import_rows(
        db,
        rows,
    )